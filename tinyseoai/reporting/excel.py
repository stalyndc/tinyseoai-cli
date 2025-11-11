from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _autofit(ws, cols: int, min_w: int = 10, max_w: int = 80):
    widths = [0] * (cols + 1)
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row, start=1):
            if cell is None:
                continue
            cell_str = str(cell)
            widths[i] = max(widths[i], len(cell_str))
    for i in range(1, cols + 1):
        w = max(min_w, min(widths[i] + 2, max_w))
        ws.column_dimensions[get_column_letter(i)].width = w


def write_xlsx(summary: Dict[str, Any], out_path: Path) -> Path:
    """
    summary is the loaded JSON from reports/<site>/summary.json
    Optionally you can merge in AI results before calling.
    """
    site = summary.get("site", "")
    pages_scanned = summary.get("pages_scanned", 0)
    issues: List[Dict[str, Any]] = summary.get("issues", [])
    meta = summary.get("meta", {})

    wb = Workbook()

    # Overview sheet
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "Site";            ws["A1"].font = Font(bold=True);     ws["B1"] = site
    ws["A2"] = "Pages scanned";   ws["A2"].font = Font(bold=True);     ws["B2"] = pages_scanned
    ws["A3"] = "Issues found";    ws["A3"].font = Font(bold=True);     ws["B3"] = len(issues)
    ws["A4"] = "Timestamp";       ws["A4"].font = Font(bold=True);     ws["B4"] = meta.get("timestamp", "")

    ai = summary.get("ai_summary")
    if ai:
        ws["A6"] = "AI summary";  ws["A6"].font = Font(bold=True)
        ws["B6"] = (ai.get("summary", "") or "")[:500]
        ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Style header col A and autofit
    for r in range(1, 7):
        ws[f"A{r}"].font = Font(bold=True)
    _autofit(ws, 2, max_w=120)  # Wider for AI summary

    # Issues sheet
    wi = wb.create_sheet("Issues")
    headers = ["Type", "Severity", "URL", "Detail"]
    wi.append(headers)
    for c in range(1, len(headers) + 1):
        wi.cell(row=1, column=c).font = Font(bold=True)
    for it in issues:
        wi.append([
            it.get("type", ""),
            it.get("severity", ""),
            it.get("url", ""),
            it.get("detail", ""),
        ])
    _autofit(wi, len(headers))

    # If AI summary is merged, add a sheet
    ai = summary.get("ai_summary")
    if ai:
        wa = wb.create_sheet("AI Summary")
        wa.append(["Field", "Value"])
        wa["A1"].font = Font(bold=True)
        wa["B1"].font = Font(bold=True)

        wa.append(["Summary", ai.get("summary", "")])
        wa.append(["Top issues (count)", len(ai.get("top_issues", []))])
        wa.append(["Actions (count)", len(ai.get("recommended_actions", []))])
        wa.append(["Plan used", ai.get("plan_used", "")])

        # Dump top issues
        wa.append([])
        wa.append(["Top issues", "Why it matters", "Evidence"])
        for cell in ("A", "B", "C"):
            wa[f"{cell}{wa.max_row}"].font = Font(bold=True)
        for item in ai.get("top_issues", []):
            wa.append([item.get("type", ""), item.get("why_it_matters", ""), item.get("evidence", "")])

        # Dump actions
        wa.append([])
        wa.append(["Recommended actions", "Impact", "Effort"])
        for cell in ("A", "B", "C"):
            wa[f"{cell}{wa.max_row}"].font = Font(bold=True)
        for step in ai.get("recommended_actions", []):
            wa.append([step.get("action", ""), step.get("impact", ""), step.get("effort", "")])

        # Wrap text
        for row in wa.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        _autofit(wa, 3, max_w=100)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
